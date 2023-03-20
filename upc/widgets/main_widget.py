import configparser as cp
import os
import threading
from datetime import datetime
import numpy as np

import cv2
import pandas as pd
import xlsxwriter
import openpyxl
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
import torch.hub
from PyQt5 import QtCore
from PyQt5.QtCore import *
from PyQt5.QtGui import *
from PyQt5.QtWidgets import *

from upc.common import general, style_sheet
from upc.common.general import LOGGER
from upc.widgets.action_widget import ActionWidget
from upc.widgets.ai_detect_widget import AiDetectWidget, Images
from upc.widgets.init_widget import InitWidget
from upc.widgets.scan_setting_widget import ScanAndSettingWidget
from upc.widgets.setting_widget import SettingWidget


class Setting:
    def __init__(self):
        super(Setting, self).__init__()

        self.iou = None
        self.conf = None
        self.top_focus = None
        self.side_focus = None
        self.rpa_qr = None
        self.yk_center = None
        self.load_config()

    def load_config(self):
        config = cp.ConfigParser()
        config.read(os.path.join(os.getcwd(), "upc", "common", "config.ini"), encoding='UTF-8')

        self.iou = config["model"]["iou"]
        self.setIou(self.iou)
        self.conf = config["model"]["conf"]
        self.setConf(self.conf)
        self.top_focus = config["camera"]["top_focus"]
        self.setTopFocus(self.top_focus)
        self.side_focus = config["camera"]["side_focus"]
        self.setSideFocus(self.side_focus)
        self.yk_center = config["yk"]["yk_center"]
        self.rpa_qr = config["yk"]["rpa_qr"]

    def save_config(self):
        config = cp.ConfigParser()
        config["model"] = {
            'iou': self.iou,
            'conf': self.conf
        }
        config["camera"] = {
            'top_focus': self.top_focus,
            'side_focus': self.side_focus
        }
        config["yk"] = {
            'yk_center': self.yk_center,
            'rpa_qr': self.rpa_qr
        }
        with open(os.path.join(os.getcwd(), "upc", "common", "config.ini"), 'w', encoding='utf-8') as f:
            config.write(f)

    def setIou(self, iou):
        self.iou = iou

    def getIou(self):
        return self.iou

    def setConf(self, conf):
        self.conf = conf

    def getConf(self):
        return self.conf

    def setTopFocus(self, top_focus):
        self.top_focus = top_focus

    def getTopFocus(self):
        return self.top_focus

    def setSideFocus(self, side_focus):
        self.side_focus = side_focus

    def getSideFocus(self):
        return self.side_focus

    def setRpaQr(self, rpa_qr):
        self.rpa_qr = rpa_qr

    def getRpaQr(self):
        return self.rpa_qr

    def setYkCenter(self, yk_center):
        self.yk_center = yk_center

    def getYkCenter(self):
        return self.yk_center


class DetectHub(QtCore.QThread):
    def __init__(self, source_yolo_path, weights, source="local", name="custom", setting=None):
        super(DetectHub, self).__init__()
        self.source_yolo_path = source_yolo_path
        self.weights = weights
        self.name = name
        self.source = source
        time = datetime.now()
        self.model_top = None
        self.model_yoko = None
        self.model_second = None
        self.model_unknown = None
        self.frame = None
        self.setting = setting
        try:
            t1 = threading.Thread(target=self.load_top_model)
            t2 = threading.Thread(target=self.load_side_model)
            t3 = threading.Thread(target=self.load_case_model)
            t1.start()
            t2.start()
            t3.start()
            t1.join()
            t2.join()
            t3.join()
        except Exception as e:
            LOGGER.error(f"LOAD MODEL FAILED : {e}")
        print(f"{datetime.now() - time}")

    def load_top_model(self):
        self.model_top = torch.hub.load(self.source_yolo_path, self.name, path=self.weights["top"], source=self.source)

    def load_side_model(self):
        self.model_yoko = torch.hub.load(self.source_yolo_path, self.name, path=self.weights["side"],
                                         source=self.source)

    def load_case_model(self):
        self.model_second = torch.hub.load(self.source_yolo_path, self.name, path=self.weights["case"],
                                           source=self.source)

    def set_inference(self):
        self.model_top.cuda()
        self.model_top.iou = float(self.setting.getIou())
        self.model_top.conf = float(self.setting.getConf())
        # self.model_top.agnostic = True

        self.model_yoko.cuda()
        self.model_yoko.iou = float(self.setting.getIou())
        self.model_yoko.conf = float(self.setting.getConf())
        # self.model_yoko.agnostic = True

        self.model_second.cuda()
        self.model_second.iou = float(self.setting.getIou())
        self.model_second.conf = float(self.setting.getConf())

    def load_unknown_model(self):
        self.model_unknown = torch.hub.load(self.source_yolo_path, self.name, path=self.weights["unknown"],
                                            source=self.source)


class Main(QMainWindow):
    notification_signal = QtCore.pyqtSignal(object)
    process_signal = QtCore.pyqtSignal(object)

    def __init__(self, *args, **kwargs):
        super(Main, self).__init__(*args, **kwargs)
        self.report_file = None
        self.setObjectName("MainWindow")
        self.dlg = None
        self.setting_widget = None
        self.ai_detect_widget = None
        self.intro_widget = None
        self.detect_hub = None
        self.action_widget = None
        self.scan_and_setting_widget = None
        self.central = None
        self.list_add_count = 0
        self.value = 0
        self.get_detected = {}
        self.setContentsMargins(0, 0, 0, 0)
        self.create_ui()
        self.setting = Setting()
        self.init()
        self.is_setting = False

    # region Create UI
    def create_ui(self):
        self.central = QStackedWidget()
        # Add Intro Widget [0]
        self.intro_widget = InitWidget()

        # Add Scan and Setting Widget [1]
        self.scan_and_setting_widget = ScanAndSettingWidget()
        try:
            self.notification_signal.connect(self.scan_and_setting_widget.notification.setText)
        except Exception as e:
            print(e)
        self.scan_and_setting_widget.btn_reload_camera.clicked.connect(self.on_reload_camera_click)
        self.scan_and_setting_widget.btn_setting.clicked.connect(self.on_btn_setting_click)

        # Add Action Widget [2]
        self.action_widget = ActionWidget()
        self.action_widget.back.clicked.connect(self.on_action_widget_btn_back_click)
        self.action_widget.action_1.clicked.connect(self.on_btn_action_1_click)

        # Add AI Detect Widget [3]
        self.ai_detect_widget = AiDetectWidget()
        self.ai_detect_widget.back.clicked.connect(self.on_ai_detect_widget_btn_back_click)
        self.ai_detect_widget.btn_detect.clicked.connect(self.on_ai_detect_widget_btn_detect_click)
        self.ai_detect_widget.btn_print.clicked.connect(self.on_ai_detect_widget_btn_detect_with_path_click)
        self.ai_detect_widget.btn_get_report.clicked.connect(self.load_data_to_sheet)
        self.ai_detect_widget.btn_open_excel.clicked.connect(self.btn_open_excel)

        # Add Setting Widget [4]
        self.setting_widget = SettingWidget()
        self.setting_widget.btn_save.clicked.connect(self.on_setting_widget_btn_save_click)

        self.central.addWidget(self.intro_widget)
        self.central.addWidget(self.scan_and_setting_widget)
        self.central.addWidget(self.action_widget)
        self.central.addWidget(self.ai_detect_widget)
        self.central.addWidget(self.setting_widget)
        self.setCentralWidget(self.central)

    # endregion
    # region Main Widget Control
    def init(self):
        # Load init widget
        t1 = threading.Thread(target=self.transfer_form, args=[])
        t1.start()

    def init_model(self):
        # Init Model
        weights = {"top": os.path.join(os.getcwd(), "assets", "models", "top.pt"),
                   "side": os.path.join(os.getcwd(), "assets", "models", "side.pt"),
                   "case": os.path.join(os.getcwd(), "assets", "models", "case.pt"),
                   "unknown": os.path.join(os.getcwd(), "assets", "models", "unknown.pt")}
        self.detect_hub = DetectHub(source_yolo_path=os.getcwd(), weights=weights, setting=self.setting)
        self.detect_hub.set_inference()

    def box2list(self, boxes):

        detected_count = {}
        for box in boxes:
            (x_min, y_min, x_max, y_max, conf, cls, name) = (
                int(box[0]), int(box[1]), int(box[2]), int(box[3]), float(box[4]), int(box[5]), str(box[6]))
            cls_name = f"{cls}_{name}"
            if cls_name in detected_count.keys():
                detected_count[cls_name] += 1
            else:
                detected_count[cls_name] = 1
            if cls_name in self.ai_detect_widget.count_total.keys():
                self.ai_detect_widget.count_total[cls_name] += 1
            else:
                self.ai_detect_widget.count_total[cls_name] = 1

        return detected_count

    def transfer_form(self):
        # try:
        #     if self.central.currentIndex() == 0:  # Init widget -> Scan Widget
        #         self.intro_widget.notification_init.setText("カメラの読み込み . . .")
        #         self.scan_and_setting_widget.load_list_camera()
        #         self.intro_widget.notification_init.setText("AI モデルの読み込み. . .")
        self.init_model()
        #
        #         self.central.setCurrentIndex(1)
        #
        #         # Check camera exist
        #         if self.scan_and_setting_widget is not None and len(self.scan_and_setting_widget.camera_list) < 2:
        #             self.notification_signal.emit("カメラが初期化されていません。カメラをチェックしてください！")
        #             return
        #         # self.notification_signal.emit("カメラが初期化されました。バーコードをスキャンしてください !")
        #
        #         # Read barcode ID
        #         while True:
        #             if not self.is_setting:
        #                 if self.scan_and_setting_widget.data_read is not None:
        #                     self.action_widget.camera_loaded = self.scan_and_setting_widget.camera_loaded
        #                     break
        #
        #         # Set Title
        #         self.action_widget.lb_center.setText(self.setting.getYkCenter())
        #         self.action_widget.rpa_qr.setText(self.setting.getRpaQr())
        #         self.action_widget.lb_yl_id.setText(f"{self.scan_and_setting_widget.data_read}")
        #         self.action_widget.lb_yl_name.setText(f"ブイニャットナム")
        # except Exception as e:
        #     LOGGER.error(e)
        #     return
        # # --> Direct to Action Widget [2]
        self.central.setCurrentIndex(2)

    # endregion
    # region Setting Widget control

    def on_setting_widget_btn_save_click(self):

        try:
            self.setting.setIou(self.setting_widget.iou.text())
            self.setting.setConf(self.setting_widget.conf.text())
            self.setting.setTopFocus(self.setting_widget.focus_top.text())
            self.setting.setSideFocus(self.setting_widget.focus_side.text())
            self.setting.setRpaQr(self.setting_widget.rpa_qr.currentText())
            self.setting.setYkCenter(self.setting_widget.yk_center.currentText())
            self.setting.save_config()
            self.detect_hub.set_inference()
            LOGGER.info("CONFIG IS SAVED")
            if self.central.currentIndex != 1:
                self.central.setCurrentIndex(1)
        except Exception as e:
            print(e)
        self.is_setting = False

    # endregion
    # region Action Widget Control

    def on_btn_action_1_click(self):
        if self.central.currentIndex != 3:
            try:
                self.ai_detect_widget.lb_center.setText(self.setting.getYkCenter())
                self.ai_detect_widget.rpa_qr.setText(self.setting.getRpaQr())
                self.ai_detect_widget.lb_yl_id.setText(f"{self.scan_and_setting_widget.data_read}")
                self.ai_detect_widget.lb_yl_name.setText(f"ブイニャットナム")
            except Exception as e:
                print(e)
            self.central.setCurrentIndex(3)

    def on_action_widget_btn_back_click(self):
        self.central.setCurrentIndex(1)
        self.scan_and_setting_widget.scan_barcode()
        t = threading.Thread(target=self.doing)
        t.start()

    def doing(self):
        while True:
            if not self.is_setting:
                if self.scan_and_setting_widget.data_read is not None:
                    break
        try:
            print("do set title lan 2")
            self.action_widget.lb_center.setText(self.setting.getYkCenter())
            self.action_widget.rpa_qr.setText(self.setting.getRpaQr())
            self.action_widget.lb_yl_name.setText(f"担当者CD：{self.scan_and_setting_widget.data_read}")
        except Exception as e:
            print(e)
        self.central.setCurrentIndex(2)

    # endregion
    # region Scan Widget Control

    def load_list_thread(self):
        is_not = False
        while True:
            if self.scan_and_setting_widget.is_loaded:
                self.scan_and_setting_widget.btn_reload_camera.setEnabled(True)
                if self.scan_and_setting_widget.camera_list is not None and len(
                        self.scan_and_setting_widget.camera_list) == 2:
                    self.notification_signal.emit("カメラが初期化されました。バーコードをスキャンしてください !")
                    break
                else:
                    self.notification_signal.emit("カメラが初期化されていません。カメラをチェックしてください！ ")
                    is_not = True
                    break
        if is_not:
            return
        while True:
            if not self.is_setting:
                if self.scan_and_setting_widget.data_read is not None:
                    self.action_widget.camera_loaded = self.scan_and_setting_widget.camera_loaded
                    break
        self.action_widget.lb_center.setText(self.setting.getYkCenter())
        self.action_widget.rpa_qr.setText(self.setting.getRpaQr())
        self.action_widget.lb_yl_id.setText(f"{self.scan_and_setting_widget.data_read}")
        self.action_widget.lb_yl_name.setText(f"ブイニャットナム")
        self.notification_signal.emit("")

        self.central.setCurrentIndex(2)

    def on_reload_camera_click(self):
        t = threading.Thread(target=self.disable_control, args=[])
        t.start()
        self.scan_and_setting_widget.data_read = None
        self.scan_and_setting_widget.load_list_camera(is_reload=True)
        t2 = threading.Thread(target=self.load_list_thread, args=[])
        t2.start()

    def on_btn_setting_click(self):
        if self.central.currentIndex != 4:
            self.is_setting = True
            self.central.setCurrentIndex(4)

    def disable_control(self):
        self.notification_signal.emit("")
        self.scan_and_setting_widget.btn_reload_camera.setEnabled(False)

    # endregion
    # region AI Detect Widget Control

    def detect(self):
        side_names = []
        with open("./assets/side.txt", "r", encoding="UTF-8") as f:
            for line in f:
                side_names.append(line.strip())
        top_names = []
        with open("./assets/top.txt", "r", encoding="UTF-8") as f:
            for line in f:
                top_names.append(line.strip())
        CASE_DIC = {'0': 13, "1": 12, "2": 14, "3": 20, "4": 21, "5": 22, "6": 23, "7": 24, "8": 25, "9": 26, "10": 27,
                    "11": 28, "12": 29, "13": 33, "14": 34, "15": 0, "16": 1, "17": 2, "18": 3, "19": 54, "20": 58,
                    "21": 57, "22": 36}
        try:
            if self.action_widget.camera_loaded is not None:
                # Get Image from camera
                self.ai_detect_widget.get_images(camera_loaded=self.action_widget.camera_loaded, setting=self.setting)
                self.process_signal.emit(25)
                QApplication.processEvents()

            while True:
                if self.ai_detect_widget.list_detecting_images is not None and len(
                        self.ai_detect_widget.list_detecting_images) == 2:
                    self.list_add_count += 1
                    for detect_image in self.ai_detect_widget.list_detecting_images:
                        # datetime.now().strftime("%Y%m%d_%H%M%S")
                        detect_image.setDescriptions(camera=detect_image.getCamera(),
                                                     time_=datetime.now().strftime("%Y/%m/%d %H.%M.%S"),
                                                     index=self.list_add_count)
                        if str(detect_image.getCamera()) == "TOP":
                            result_top = self.detect_hub.model_top(detect_image.getImage())
                            boxes = []
                            if result_top is not None:
                                # TODO set to box

                                draw_img = detect_image.getDrawImage()
                                for box in result_top.xyxy[0]:
                                    (case_x_min, case_y_min, case_x_max, case_y_max, case_conf, case_cls) = (
                                        int(box[0]), int(box[1]), int(box[2]), int(box[3]), float(box[4]), int(box[5]))
                                    temp_box = (
                                        case_x_min, case_y_min, case_x_max, case_y_max, case_conf, case_cls,
                                        top_names[case_cls])

                                    boxes.append(temp_box)
                                # TODO Remove over box in boxes

                                for box in boxes:
                                    draw_img = self.draw_put_text(draw_img, int(box[0]), int(box[1]), int(box[2]),
                                                                  int(box[3]), float(box[4]), int(box[5]), point=True)

                                detect_image.setDetectedBBox(boxes)
                                detect_image.setDetectedCount(self.box2list(boxes))
                                detect_image.setDrawImage(draw_img)

                        elif str(detect_image.getCamera()) == "SIDE":
                            result_side = self.detect_hub.model_yoko(detect_image.getImage())

                            boxes = []
                            if result_side is not None:

                                draw_img = detect_image.getDrawImage()

                                # Add detected box to boxes
                                for box in result_side.xyxy[0]:
                                    (case_x_min, case_y_min, case_x_max, case_y_max, case_conf, case_cls) = (
                                        int(box[0]), int(box[1]), int(box[2]), int(box[3]), float(box[4]), int(box[5]))

                                    self.detect_hub.model_second.iou = float(self.setting.getIou())
                                    self.detect_hub.model_second.conf = float(self.setting.getConf())
                                    if case_cls in [13, 12, 14, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 33, 34, 0, 1, 2,
                                                    3, 54, 58, 57, 36]:  # アロエ, 鉄
                                        input_image = detect_image.getImage()[case_y_min:case_y_max,
                                                      case_x_min:case_x_max]
                                        input_image1 = cv2.cvtColor(input_image, cv2.COLOR_BGR2RGB)
                                        # now = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                                        # cv2.imwrite(f"{now}.png", input_image1)

                                        result_case = self.detect_hub.model_second(input_image)
                                        xyxy_ = result_case.xyxy[0]
                                        if xyxy_ is not None:
                                            # max_box=max(xyxy_, key=lambda box : box[4])
                                            for box_ in xyxy_:
                                                (x1, y1, x2, y2, conf_1, cls_1) = (
                                                    int(box_[0]), int(box_[1]), int(box_[2]), int(box_[3]),
                                                    float(box_[4]), int(box_[5]))
                                                # cv2.rectangle(input_image1, (x1, y1), (x2, y2), (0, 255, 255), 1)
                                                input_image1 = cv2.putText(input_image1, f"{cls_1}",
                                                                           (x1, y1),
                                                                           cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 1,
                                                                           cv2.LINE_4)
                                                now1 = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                                                cv2.imwrite(f"{now1}.png", input_image1)
                                                print(f"tim ra thang {cls_1} {case_cls}=>> {cls_1}")
                                                if int(case_cls) == int(CASE_DIC[str(cls_1)]):
                                                    case_conf = 0.99
                                                    case_cls = CASE_DIC[str(cls_1)]
                                                else:
                                                    if case_conf < 0.7:
                                                        pass
                                                    else:
                                                        case_cls = CASE_DIC[str(cls_1)]
                                                        case_conf = conf_1

                                    temp_box = (
                                        case_x_min, case_y_min, case_x_max, case_y_max, case_conf, case_cls,
                                        side_names[case_cls])
                                    boxes.append(temp_box)
                                # TODO Remove overlap box in boxes
                                # sorted_boxes = sorted(boxes, key=lambda cls: cls[5])

                                for box in boxes:
                                    draw_img = self.draw_put_text(draw_img, int(box[0]), int(box[1]), int(box[2]),
                                                                  int(box[3]), float(box[4]), int(box[5]), point=True)

                                detect_image.setDrawImage(draw_img)
                                detect_image.setDetectedBBox(boxes)
                                detect_image.setDetectedCount(self.box2list(boxes))

                        # Add detected object to total list
                        self.ai_detect_widget.list_detected_images_total.append(detect_image)
                        self.process_signal.emit(50)

                    # Clear list detecting image
                    self.ai_detect_widget.update_list_image()
                    self.ai_detect_widget.update_table_view()
                    self.ai_detect_widget.list_detecting_images = []
                    self.process_signal.emit(75)
                    QApplication.processEvents()
                    break
        except Exception as e:
            print(e)

    def detect_with_path(self, is_create_anno=False, is_draw_point=False, is_not_show_conf=False):
        side_names = []
        with open("./assets/side.txt", "r", encoding="UTF-8") as f:
            for line in f:
                side_names.append(line.strip())
        top_names = []
        with open("./assets/top.txt", "r", encoding="UTF-8") as f:
            for line in f:
                top_names.append(line.strip())
        CASE_DIC = {'0': 13, "1": 12, "2": 14, "3": 20, "4": 21, "5": 22, "6": 23, "7": 24, "8": 25, "9": 26, "10": 27,
                    "11": 28, "12": 29, "13": 33, "14": 34, "15": 0, "16": 1, "17": 2, "18": 3, "19": 54, "20": 58,
                    "21": 57, "22": 36}
        try:
            list_image = {}
            if len(self.ai_detect_widget.txt_folder.text()) > 0:
                file, _ = general.get_all_file(self.ai_detect_widget.txt_folder.text())
            else:
                print("rong")
                return
            for img in file:
                if img.endswith("png") or img.endswith("jpg"):
                    frame = cv2.imread(img)
                    frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
                    image = Images()
                    image.setImage(frame.copy())
                    image.setDrawImage(frame.copy())
                    if "top" in self.ai_detect_widget.txt_folder.text():
                        image.setCamera("TOP")
                    elif "side" in self.ai_detect_widget.txt_folder.text():
                        image.setCamera("SIDE")
                    else:
                        image.setCamera("CASE")
                    list_image[os.path.basename(img)] = image
                    self.get_detected[os.path.basename(img)] = []

            for image_name, detect_image in list_image.items():
                self.process_signal.emit(50)
                # datetime.now().strftime("%Y%m%d_%H%M%S")
                detect_image.setDescriptions(camera=f"{image_name} ({detect_image.getCamera()})",
                                             time_=datetime.now().strftime("%Y/%m/%d %H.%M.%S"),
                                             index=self.list_add_count)
                if str(detect_image.getCamera()) == "TOP":
                    result_top = self.detect_hub.model_top(detect_image.getImage())
                    boxes = []
                    if result_top is not None:
                        # TODO set to box

                        draw_img = detect_image.getDrawImage()
                        for box in result_top.xyxy[0]:
                            (case_x_min, case_y_min, case_x_max, case_y_max, case_conf, case_cls) = (
                                int(box[0]), int(box[1]), int(box[2]), int(box[3]), float(box[4]), int(box[5]))
                            temp_box = (
                                case_x_min, case_y_min, case_x_max, case_y_max, case_conf, case_cls,
                                top_names[case_cls])
                            yolo_box = self.bnd_box_to_yolo_line((case_x_min, case_y_min, case_x_max, case_y_max),
                                                                 draw_img.shape[:2])
                            if is_create_anno:
                                save_txt = self.ai_detect_widget.txt_folder.text()
                                if len(save_txt) > 0:
                                    with open(os.path.join(save_txt, image_name[:-4] + ".txt"), "a",
                                              encoding="utf-8") as f:
                                        f.write(f"{case_cls} {yolo_box[0]} {yolo_box[1]} {yolo_box[2]} {yolo_box[3]}\n")
                            boxes.append(temp_box)
                        self.get_detected[image_name] = boxes

                        # TODO Remove over box in boxes

                        for box in boxes:
                            draw_img = self.draw_put_text(draw_img, int(box[0]), int(box[1]), int(box[2]),
                                                          int(box[3]), float(box[4]), int(box[5]), point=is_draw_point,
                                                          not_show_conf=is_not_show_conf)

                        detect_image.setDetectedBBox(boxes)
                        detect_image.setDetectedCount(self.box2list(boxes))
                        detect_image.setDrawImage(draw_img)
                        # cv2.imwrite(f"{image_name}", cv2.cvtColor(detect_image.getDrawImage(), cv2.COLOR_BGR2RGB))

                elif str(detect_image.getCamera()) == "SIDE":
                    result_side = self.detect_hub.model_yoko(detect_image.getImage())
                    boxes = []
                    if result_side is not None:

                        draw_img = detect_image.getDrawImage()

                        # Add detected box to boxes
                        for box in result_side.xyxy[0]:
                            (case_x_min, case_y_min, case_x_max, case_y_max, case_conf, case_cls) = (
                                int(box[0]), int(box[1]), int(box[2]), int(box[3]), float(box[4]), int(box[5]))
                            #
                            if case_cls in [13, 12, 14, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 33, 34, 0, 1, 2, 3, 54,
                                            58, 57, 36]:  # アロエ, 鉄
                                input_image = detect_image.getImage()[case_y_min:case_y_max, case_x_min:case_x_max]
                                # input_image1 = cv2.cvtColor(input_image, cv2.COLOR_BGR2RGB)
                                # now = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                                # cv2.imwrite(f"{now}.png", input_image1)

                                result_case = self.detect_hub.model_second(input_image)
                                xyxy_ = result_case.xyxy[0]
                                if xyxy_ is not None:
                                    # max_box=max(xyxy_, key=lambda box : box[4])
                                    for box_ in xyxy_:
                                        (x1, y1, x2, y2, conf_1, cls_1) = (
                                            int(box_[0]), int(box_[1]), int(box_[2]), int(box_[3]),
                                            float(box_[4]), int(box_[5]))
                                        # cv2.rectangle(input_image1, (x1, y1), (x2, y2), (0, 255, 255), 1)
                                        # input_image1 = cv2.putText(input_image1, f"{cls_1}",
                                        #                            (x1, y1),
                                        #                            cv2.FONT_HERSHEY_SIMPLEX, 1, (255, 0, 0), 1,
                                        #                            cv2.LINE_4)
                                        # now1 = datetime.now().strftime("%Y%m%d_%H%M%S_%f")
                                        # cv2.imwrite(f"{now1}.png", input_image1)
                                        print(f"FIRST : {case_cls}=>> SECOND : {CASE_DIC[str(cls_1)]}")
                                        if int(case_cls) == int(CASE_DIC[str(cls_1)]):
                                            case_conf = 0.99
                                            case_cls = CASE_DIC[str(cls_1)]
                                        else:
                                            if conf_1 > 0.7:
                                                case_cls = CASE_DIC[str(cls_1)]
                                                case_conf = conf_1

                            temp_box = (
                                case_x_min, case_y_min, case_x_max, case_y_max, case_conf, case_cls,
                                side_names[case_cls])
                            boxes.append(temp_box)
                            yolo_box = self.bnd_box_to_yolo_line((case_x_min, case_y_min, case_x_max, case_y_max),
                                                                 draw_img.shape[:2])
                            # create annotation for image
                            if is_create_anno:
                                save_txt = self.ai_detect_widget.txt_folder.text()
                                if len(save_txt) > 0:
                                    with open(os.path.join(save_txt, image_name[:-4] + ".txt"), "a",
                                              encoding="utf-8") as f:
                                        f.write(f"{case_cls} {yolo_box[0]} {yolo_box[1]} {yolo_box[2]} {yolo_box[3]}\n")
                        # TODO Remove overlap box in boxes
                        # sorted_boxes = sorted(boxes, key=lambda cls: cls[5])
                        self.get_detected[image_name] = boxes
                        for box in boxes:
                            draw_img = self.draw_put_text(draw_img,
                                                          int(box[0]),  # x1
                                                          int(box[1]),  # y1
                                                          int(box[2]),  # x2
                                                          int(box[3]),  # y2
                                                          float(box[4]),  # confidence
                                                          int(box[5]),  # cls
                                                          point=is_draw_point,
                                                          not_show_conf=is_not_show_conf)  # draw point

                        detect_image.setDrawImage(draw_img)
                        detect_image.setDetectedBBox(boxes)
                        detect_image.setDetectedCount(self.box2list(boxes))
                elif str(detect_image.getCamera()) == "CASE":
                    result_case = self.detect_hub.model_second(detect_image.getImage())
                    self.detect_hub.model_second.iou = float(self.setting.getIou())
                    self.detect_hub.model_second.conf = float(self.setting.getConf())
                    self.detect_hub.model_second.multi_label = False
                    # self.detect_hub.model_top.agnostic = True
                    boxes = []
                    if result_case is not None:
                        # TODO set to box

                        draw_img = detect_image.getDrawImage()
                        for box in result_case.xyxy[0]:
                            (case_x_min, case_y_min, case_x_max, case_y_max, case_conf, case_cls) = (
                                int(box[0]), int(box[1]), int(box[2]), int(box[3]), float(box[4]), int(box[5]))
                            temp_box = (
                                case_x_min, case_y_min, case_x_max, case_y_max, case_conf, case_cls,
                                top_names[case_cls])
                            yolo_box = self.bnd_box_to_yolo_line((case_x_min, case_y_min, case_x_max, case_y_max),
                                                                 draw_img.shape[:2])
                            if is_create_anno:
                                save_txt = self.ai_detect_widget.txt_folder.text()
                                if len(save_txt) > 0:
                                    with open(os.path.join(save_txt, image_name[:-4] + ".txt"), "a",
                                              encoding="utf-8") as f:
                                        f.write(f"{case_cls} {yolo_box[0]} {yolo_box[1]} {yolo_box[2]} {yolo_box[3]}\n")
                            boxes.append(temp_box)
                        self.get_detected[image_name] = boxes

                        # TODO Remove over box in boxes

                        for box in boxes:
                            draw_img = self.draw_put_text(draw_img, int(box[0]), int(box[1]), int(box[2]),
                                                          int(box[3]), float(box[4]), int(box[5]), point=is_draw_point,
                                                          not_show_conf=is_not_show_conf)

                        detect_image.setDetectedBBox(boxes)
                        detect_image.setDetectedCount(self.box2list(boxes))
                        detect_image.setDrawImage(draw_img)
                        # cv2.imwrite(f"{image_name}", cv2.cvtColor(detect_image.getDrawImage(), cv2.COLOR_BGR2RGB))

                # Add detected object to total list
                self.ai_detect_widget.list_detected_images_total.append(detect_image)

                # Clear list detecting image
            self.process_signal.emit(75)
            QApplication.processEvents()
            self.ai_detect_widget.update_list_image(list_image=list_image.values())
            self.ai_detect_widget.update_table_view()
            self.ai_detect_widget.list_detecting_images = []
            self.process_signal.emit(100)
            QApplication.processEvents()
        except Exception as e:
            print(e)

    @staticmethod
    def bnd_box_to_yolo_line(box, img_size):
        x_min, y_min = box[0], box[1]
        x_max, y_max = box[2], box[3]

        x_center = float((x_min + x_max)) / 2 / img_size[1]
        y_center = float((y_min + y_max)) / 2 / img_size[0]

        w = float((x_max - x_min)) / img_size[1]
        h = float((y_max - y_min)) / img_size[0]

        return x_center, y_center, w if w < 1 else 1.0, h if h < 1 else 1.0

    @staticmethod
    def draw_put_text(draw_img, x1, y1, x2, y2, conf, cls, point=False, not_show_conf=False):

        if int(conf * 100) >= 80:
            color = (0, 0, 255)  # BLUE
        elif int(conf * 100) >= 50:
            color = (0, 255, 0)  # GREEN
        else:
            color = (255, 0, 0)  # RED
        if point:
            # over = draw_img.copy()
            # cv2.circle(over, (int(x1 + (x2 - x1) / 2), int(y1 + (y2 - y1) / 2)), 28, color, thickness=-1,
            #            lineType=cv2.LINE_4, shift=0)
            # alpha = 0.4
            #
            # draw_img = cv2.addWeighted(over, alpha, draw_img, 1 - alpha, 0)

            # cv2.circle(draw_img, (int(x1 + (x2 - x1) / 2), int(y1 + (y2 - y1) / 2)), 30, color, thickness=5,
            #            lineType=cv2.LINE_4,
            #            shift=0)
            cv2.ellipse(draw_img, center=(int(x1 + (x2 - x1) / 2), int(y1 + (y2 - y1) / 2)),
                        axes=(40, 40),
                        angle=270,
                        startAngle=0,
                        endAngle=360,
                        color=(128, 128, 128),
                        thickness=7,
                        lineType=cv2.LINE_4,
                        shift=0)
            cv2.ellipse(draw_img, center=(int(x1 + (x2 - x1) / 2), int(y1 + (y2 - y1) / 2)),
                        axes=(40, 40),
                        angle=270,
                        startAngle=0,
                        endAngle=int(conf * 360),
                        color=color,
                        thickness=7,
                        lineType=cv2.LINE_4,
                        shift=0)
            cv2.putText(draw_img, f"{cls}",
                        (int(x1 + (x2 - x1) / 2) - 20, int(y1 + (y2 - y1) / 2) + 10),
                        cv2.FONT_HERSHEY_COMPLEX, 1,
                        (255, 255, 255), 6, cv2.LINE_AA)
            cv2.putText(draw_img, f"{cls}",
                        (int(x1 + (x2 - x1) / 2) - 20, int(y1 + (y2 - y1) / 2) + 10),
                        cv2.FONT_HERSHEY_COMPLEX, 1, color,
                        2, cv2.LINE_AA)
        else:
            # cv2.rectangle(draw_img, (x1, y1),
            #               (x2, y2), color, 2)
            # cv2.putText(draw_img, f"{cls} {int(conf * 100)}%", (x1 + 10, y1 + 30), cv2.FONT_HERSHEY_COMPLEX, 1,
            #             (255, 255, 255), 6, cv2.LINE_AA)
            # cv2.putText(draw_img, f"{cls} {int(conf * 100)}%", (x1 + 10, y1 + 30), cv2.FONT_HERSHEY_COMPLEX, 1, color,
            #             3, cv2.LINE_AA)

            z = 30
            x = 10
            cx, cy = (x1 + x2) // 2, (y1 + y2) // 2
            if (x2 - x1) < (y2 - y1):
                z = (cx - x1) // 2
                x = z // 3
            else:
                z = (cy - y1) // 2
                x = z // 3

            thickness = 2
            if not_show_conf:
                text = f"{cls}"
            else:
                text = f"{cls} {int(conf * 100)}%"
            if cls >= 10:
                text_pos = (int(cx - z // 2), int(cy + z // 2))
            else:
                text_pos = (int(cx - z // 3), int(cy + z // 3))

            cv2.putText(draw_img, text, text_pos, cv2.FONT_HERSHEY_SIMPLEX, 1,
                        (255, 255, 255), 7, cv2.LINE_AA)
            cv2.putText(draw_img, text, text_pos, cv2.FONT_HERSHEY_SIMPLEX, 1, color,
                        3, cv2.LINE_AA)
            color1 = (255, 255, 255)
            color = (0, 255, 0)
            thickness_border = 3
            pts1 = np.array([[cx - z, cy - z + x], [cx - z, cy - z], [cx - z + x, cy - z]], np.int32)
            pts1 = pts1.reshape((-1, 1, 2))
            draw_img = cv2.polylines(draw_img, pts=[pts1], isClosed=False, color=color1,
                                     thickness=thickness + thickness_border)
            draw_img = cv2.polylines(draw_img, pts=[pts1], isClosed=False, color=color, thickness=thickness)

            pts1 = np.array([[cx + z, cy - z + x], [cx + z, cy - z], [cx + z - x, cy - z]], np.int32)
            pts1 = pts1.reshape((-1, 1, 2))
            draw_img = cv2.polylines(draw_img, pts=[pts1], isClosed=False, color=color1,
                                     thickness=thickness + thickness_border)
            draw_img = cv2.polylines(draw_img, pts=[pts1], isClosed=False, color=color, thickness=thickness)

            pts1 = np.array([[cx - z, cy + z - x], [cx - z, cy + z], [cx - z + x, cy + z]], np.int32)
            pts1 = pts1.reshape((-1, 1, 2))
            draw_img = cv2.polylines(draw_img, pts=[pts1], isClosed=False, color=color1,
                                     thickness=thickness + thickness_border)
            draw_img = cv2.polylines(draw_img, pts=[pts1], isClosed=False, color=color, thickness=thickness)

            pts1 = np.array([[cx + z, cy + z - x], [cx + z, cy + z], [cx + z - x, cy + z]], np.int32)
            pts1 = pts1.reshape((-1, 1, 2))
            draw_img = cv2.polylines(draw_img, pts=[pts1], isClosed=False, color=color1,
                                     thickness=thickness + thickness_border)
            draw_img = cv2.polylines(draw_img, pts=[pts1], isClosed=False, color=color, thickness=thickness)

        return draw_img

    def on_ai_detect_widget_btn_back_click(self):
        self.ai_detect_widget.clear_layout()
        self.list_add_count = 0
        self.central.setCurrentIndex(2)

    def on_ai_detect_widget_btn_detect_click(self):
        try:
            self.load()
        except Exception as e:
            print(e)

    def on_ai_detect_widget_btn_detect_with_path_click(self):
        try:
            self.load_with_path()
        except Exception as e:
            print(e)

    def load(self):
        self.dlg = QProgressDialog(None, None, 0, 100, self)
        self.dlg.setStyleSheet(style_sheet.QProgressBartStyle())
        self.dlg.setWindowFlag(Qt.FramelessWindowHint)
        self.dlg.setWindowModality(Qt.WindowModal)
        self.dlg.setAttribute(QtCore.Qt.WA_TranslucentBackground, True)
        self.dlg.setAutoClose(True)
        self.dlg.show()
        self.process_signal.connect(self.dlg.setValue)
        self.process_signal.emit(0)
        QApplication.processEvents()
        t1 = datetime.now()
        self.detect()
        t2 = datetime.now() - t1
        print(t2)
        self.process_signal.emit(100)
        QApplication.processEvents()
        del self.dlg

    def load_data_to_sheet(self):
        self.dlg = QProgressDialog(None, None, 0, 100, self)
        self.dlg.setStyleSheet(style_sheet.QProgressBartStyle())
        self.dlg.setWindowFlag(Qt.FramelessWindowHint)
        self.dlg.setWindowModality(Qt.WindowModal)
        self.dlg.setAttribute(QtCore.Qt.WA_TranslucentBackground, True)
        self.dlg.setAutoClose(True)
        self.process_signal.connect(self.dlg.setValue)
        self.process_signal.emit(0)
        self.dlg.show()
        QApplication.processEvents()
        t1 = datetime.now()
        self.add_to_sheet()
        t2 = datetime.now() - t1
        print(t2)
        self.process_signal.emit(100)
        QApplication.processEvents()
        del self.dlg

    def load_with_path(self):
        self.dlg = QProgressDialog(None, None, 0, 100, self)
        self.dlg.setStyleSheet(style_sheet.QProgressBartStyle())
        self.dlg.setWindowFlag(Qt.FramelessWindowHint)
        self.dlg.setWindowModality(Qt.WindowModal)
        self.dlg.setAttribute(QtCore.Qt.WA_TranslucentBackground, True)
        self.dlg.setAutoClose(True)
        self.dlg.show()
        self.process_signal.connect(self.dlg.setValue)
        self.process_signal.emit(0)
        QApplication.processEvents()

        t1 = datetime.now()
        is_anno = False
        is_point = False
        is_not_show_conf = False
        if self.ai_detect_widget.cb_create_annotation.isChecked():
            is_anno = True
        if self.ai_detect_widget.cb_draw_point.isChecked():
            is_point = True
        if self.ai_detect_widget.cb_not_show_conf.isChecked():
            is_not_show_conf = True

        self.detect_with_path(is_create_anno=is_anno, is_draw_point=is_point, is_not_show_conf=is_not_show_conf)
        t2 = datetime.now() - t1
        print(t2)
        QApplication.processEvents()
        del self.dlg
        self.process_signal.emit(100)

    # endregion
    def add_to_sheet(self):
        self.ai_detect_widget.btn_get_report.setEnabled(False)
        if len(self.ai_detect_widget.txt_folder.text()) == 0:
            return
        if len(self.get_detected.keys()) == 0:
            return
        LOGGER.info("ADD REPORT TO EXCEL ")
        if "side" not in self.ai_detect_widget.txt_folder.text():
            files_name = './assets/top_report.xlsx'
        else:
            files_name = './assets/side_report.xlsx'

        # Init Writer
        writer = pd.ExcelWriter(files_name, engine='openpyxl', mode='a', if_sheet_exists='replace')

        # Get all file test images
        files, _ = general.get_all_file(self.ai_detect_widget.txt_folder.text())
        files = sorted(files)

        # Add data to DataFrame
        original_anno_frames = self.original_annotation_dataframe(files)
        df = pd.DataFrame(original_anno_frames)
        df.to_excel(writer, sheet_name="Sheet1", index=False)
        LOGGER.info("ADD ORIGINAL ANNOTATION TO Sheet1 SUCCESS")
        self.process_signal.emit(25)

        detected_annotation_frames = self.detected_annotation_dataframe(self, files)
        df = pd.DataFrame(detected_annotation_frames)
        df.to_excel(writer, sheet_name="Sheet2", index=False)
        LOGGER.info("ADD DETECTED ANNOTATION TO Sheet2 SUCCESS")
        self.process_signal.emit(50)
        writer.save()
        self.process_signal.emit(75)
        list_dif = {}
        try:
            # Get different result
            for (k, v), (k1, v1) in zip(original_anno_frames.items(), detected_annotation_frames.items()):
                for index, (x, y) in enumerate(zip(v, v1)):
                    if x != y:
                        if original_anno_frames["Images"][index] not in list_dif.keys():
                            list_dif[original_anno_frames["Images"][index]] = [(k, x, y)]
                        else:
                            list_dif[original_anno_frames["Images"][index]].append((k, x, y))
            LOGGER.info("ADD different result to dic")

            for image_ in list_dif.keys():
                for index, image in enumerate(self.ai_detect_widget.list_detected_images_total):
                    if image.getDescriptions()["CAMERA"].split(" ")[0] == image_:
                        save = os.path.join(os.getcwd(), "assets", "failed_images", image_)
                        save_img = cv2.cvtColor(image.getDrawImage(), cv2.COLOR_BGR2RGB)
                        h, w = save_img.shape[:2]
                        ratio = float(960 / w)
                        save_img = cv2.resize(save_img, (0, 0), fx=ratio, fy=ratio)
                        cv2.imwrite(save, save_img)
                        break
            LOGGER.info("FAILED IMAGE was saved")
            images_, _ = general.get_all_file(os.path.join(os.getcwd(), "assets", "failed_images"))

            wb = openpyxl.load_workbook(files_name)
            wb.create_sheet("Sheet4")
            ws = wb.get_sheet_by_name("Sheet4")

            idx, col = 1, 17
            for k, v in list_dif.items():
                save = os.path.join(os.getcwd(), "assets", "failed_images", k)
                ws.add_image(Image(save), anchor='A' + str(idx))
                ws.cell(row=idx, column=col, value=os.path.basename(k))

                for col_idx, i in enumerate(v):
                    ws.cell(row=idx + 1, column=col + col_idx, value=f"ID_{i[0]}")
                    ws.cell(row=idx + 2, column=col + col_idx, value=f"{i[1]} ➡ {i[2]}")
                    pass
                idx += 30

            now = datetime.now().strftime("%Y%m%d_%H%M%S")
            self.report_file = f"{files_name[:-5]}_{now}.xlsx"
            wb.save(self.report_file)
            for im in images_:
                os.remove(im)
            LOGGER.info(" REPORT SAVED")
        except Exception as e:
            print(e)

    @staticmethod
    def original_annotation_dataframe(files):
        data_frame, images, classes, index = {}, [], {}, 0

        # Get original annotation
        for f in files:
            if f.endswith("png") or f.endswith("jpg"):
                images.append(os.path.basename(f))
                data = {}
                for index in range(0, 1000, 1):
                    data[index] = 0
                txt_file = f[:-4] + ".txt"
                if os.path.isfile(txt_file):
                    with open(txt_file, "r") as txt:
                        for line in txt:
                            texts = line.strip().split(" ")
                            data[int(texts[0])] += 1
                    # my_data = {int(k): v for k, v in data.items()}
                    classes[os.path.basename(f)] = data

        name = []
        data_frame["Images"] = []
        for i in range(0, 1000, 1):
            data_frame[i] = []

        for k2, v2 in classes.items():
            name.append(k2)
            data_frame["Images"] = name
            for k, v in v2.items():
                data_frame[k].append(v)
        return data_frame

    @staticmethod
    def detected_annotation_dataframe(self, files):
        data_frame, images, classes, index = {}, [], {}, 0
        # Get Detected annotation
        for f in files:
            if f.endswith("png") or f.endswith("jpg"):
                data = {}
                for index in range(0, 1000, 1):
                    data[index] = 0
                txt_file = f[:-4] + ".txt"
                if os.path.isfile(txt_file):
                    for image_name, boxes in self.get_detected.items():
                        if os.path.basename(f) == image_name:
                            for box in boxes:
                                if int(box[5]) in data.keys():
                                    data[int(box[5])] += 1
                    classes[os.path.basename(f)] = data
        name = []
        data_frame["Images"] = []
        for i in range(0, 1000, 1):
            data_frame[i] = []

        for k2, v2 in classes.items():
            name.append(k2)
            data_frame["Images"] = name
            for k, v in v2.items():
                data_frame[k].append(v)
        return data_frame

    def btn_open_excel(self):
        if self.report_file is not None:
            os.system(f'start "excel" "{self.report_file}"')
